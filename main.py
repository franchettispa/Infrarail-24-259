# versione funzionante con colori, crea una nuova sheet per ogni famiglia di elementi analizzata e crea summary finale
# libraries
import pandas as  pd
from difflib import SequenceMatcher
import os
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# function to match most similar str from list and return percentage of match
def get_string_similarity(a, b):
    # Normalize strings
    a = a.lower().strip()
    b = b.lower().strip()
    # Remove special characters
    a = ''.join(e for e in a if e.isalnum() or e.isspace())
    b = ''.join(e for e in b if e.isalnum() or e.isspace())
    return SequenceMatcher(None, a, b).ratio()

# Clean column names while preserving GUID
def clean_column_name(col):
    if col == 'GUID':
        return col
    return col.split('\n')[0].strip()

# functino to match family to element in reference data
def match_famiglia_to_elemento(df_reference, df_exported):
    """
    Match Famiglia values to ELEMENTO values using string similarity
    
    Parameters:
    df_reference: DataFrame containing reference data with ELEMENTO column
    df_exported: DataFrame containing exported data with Famiglia column
    
    Returns:
    Dictionary of matches {famiglia: best_matching_elemento} or None if no Famiglia column
    """
    # Check if Famiglia column exists
    if 'Famiglia' not in df_exported.columns:
        return None
        
    # Get unique ELEMENTI from reference
    elementi_list = df_reference['ELEMENTO'].unique().tolist()
    
    # Get unique Famiglia value from exported data
    famiglia = df_exported['Famiglia'].iloc[0]  # Assuming same value in all rows
    
    # Find best match
    best_match = max(elementi_list, 
                    key=lambda x: get_string_similarity(famiglia, x))
    
    similarity = get_string_similarity(famiglia, best_match)
    
    return {
        'Famiglia': famiglia,
        'Matched_Elemento': best_match,
        'Similarity_Score': similarity
    }
    
# checks after str cleaning if the parameters required (from df_reference matched with element) are present if not prints what it is missing
def check_parameters(required_params, actual_params):
    
    # Normalize strings in both lists - convert to lowercase and strip whitespace
    required_set = {str(p).lower().strip() for p in required_params}
    actual_set = {str(p).lower().strip() for p in actual_params}

    # Check if all required parameters are present
    missing_params = required_set - actual_set
    all_present = len(missing_params) == 0

    return {
        'all_parameters_present': all_present,
        'missing_parameters': sorted(list(missing_params))
    }
    
# Clean column names while preserving GUID
def clean_column_name(col):
    if col == 'GUID':
        return col
    return col.split('\n')[0].strip()

# imports reference data from excel file and cleans it (removes some columns and rows thought to not be needed)
def import_reference_data(file_path, sheet_name='Modello_Dati'):
    """
    Import and clean reference data from Excel file.
    
    Parameters:
        file_path (str): Path to reference Excel file
        sheet_name (str): Name of sheet containing reference data
    
    Returns:
        pandas.DataFrame: Cleaned reference data
    """
    try:
        # Read reference Excel file
        df_reference = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=4)
        print(f"Successfully loaded reference data from {file_path}")
        # Drop empty rows from the PE column
        df_reference = df_reference.dropna(subset=['PE'])
        # drop columns not needed drom list
        to_drop = ['Unnamed: 0','PROGETTO','FAMIGLIA','DOCFAP','PFTE','PED','PE', 'ASB']
        #dropping the columns
        df_reference = df_reference.drop(columns=to_drop)
        return df_reference
        
    except Exception as e:
        print(f"Error loading reference data: {str(e)}")
        return None

# imports to be tested data from excel file and cleans it
def import_test_data(file_path):
    """
    Test function for importing and cleaning Excel data.
    
    Parameters:
        file_path (str): Path to test Excel file
    
    Returns:
        dict: Test results with processed dataframes and validation info
    """
    try:
        # Validate file existence
        if not os.path.exists(file_path):
            return {'success': False, 'error': f'File not found: {file_path}'}

        # Load Excel file
        full_excel = pd.ExcelFile(file_path)
        sheet_names = full_excel.sheet_names
        df_creati = []
        test_results = {
            'success': True,
            'processed_sheets': [],
            'skipped_sheets': [],
            'errors': [],
            'dataframes': []
        }

        for sheet in sheet_names:
            try:
                # Skip specific sheets
                if sheet in ['Instructions', 'ParamValues']:
                    test_results['skipped_sheets'].append(sheet)
                    continue

                # Determine rows to skip
                skip_number = 2 if sheet == 'Zone riscaldamento, ventilazion' else 1

                # Process sheet
                df = full_excel.parse(sheet, skiprows=skip_number)
                
                # Validate dataframe
                if df.empty:
                    test_results['errors'].append(f'Empty dataframe in sheet: {sheet}')
                    continue

                # Clean columns
                df.columns = [clean_column_name(col) for col in df.columns]
                df = df.dropna(axis=1, how='all')

                # Store results
                test_results['processed_sheets'].append(sheet)
                test_results['dataframes'].append(df)

            except Exception as e:
                test_results['errors'].append(f'Error in sheet {sheet}: {str(e)}')

        return test_results

    except Exception as e:
        return {'success': False, 'error': str(e)}

# function that writes the checked parameters to an excel file
def check_df_parameters(df_reference, df_export, writer, missing_params_summary):
    """Check parameters for a single dataframe and write to shared Excel writer."""
    match = match_famiglia_to_elemento(df_reference, df_export) # function to match famiglia to elemento returns a dict or None
    if not match:
        return "Skipping dataframe - No 'Famiglia' column found"
    
    famiglia = match['Famiglia']
    elemento = match['Matched_Elemento']
    
    df_elemento = df_reference[df_reference['ELEMENTO'] == elemento]
    param_current = list(df_export.columns)
    
    df_comparison = df_elemento.copy()
    df_comparison['parameter_exists'] = df_comparison['PARAMETRI INFORMATIVI'].isin(param_current)
    
    param_mapping = {}
    for ref_param in df_comparison['PARAMETRI INFORMATIVI']:
        for curr_param in param_current:
            if ref_param.lower().strip() == curr_param.lower().strip():
                param_mapping[ref_param] = curr_param
                break
    
    df_comparison['param_current'] = df_comparison['PARAMETRI INFORMATIVI'].map(param_mapping)
    
    # Track missing parameters for summary
    missing_params = df_comparison[~df_comparison['parameter_exists']]
    if not missing_params.empty:
        missing_params_summary.append({
            'Famiglia': famiglia,
            'Elemento': elemento,
            'Missing Parameters': ', '.join(missing_params['PARAMETRI INFORMATIVI'].tolist())
        })
    
    # Write to sheet named after famiglia
    sheet_name = f"{famiglia[:31]}"  # Excel sheet names limited to 31 chars
    df_comparison.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Format sheet
    worksheet = writer.sheets[sheet_name]
    # Auto-adjust columns
    for idx, col in enumerate(df_comparison.columns, start=1):
        max_length = max(df_comparison[col].astype(str).apply(len).max(), len(col))
        worksheet.column_dimensions[get_column_letter(idx)].width = max_length + 2
    
    # Apply colors to parameter_exists column
    param_exists_index = list(df_comparison.columns).index('parameter_exists') + 1
    col_letter = get_column_letter(param_exists_index)
    
    for row in range(2, len(df_comparison) + 2):
        cell = worksheet[f'{col_letter}{row}']
        if df_comparison.iloc[row - 2]['parameter_exists']:
            cell.fill = green_fill
        else:
            cell.fill = red_fill
            
    return f"Processed {famiglia}"

# function to create a summary sheet with missing parameters
def create_summary_sheet(writer, missing_params_summary):
    """
    Creates and formats a summary sheet in the Excel workbook showing missing parameters.
    
    Args:
        writer: ExcelWriter object
        missing_params_summary: List of dicts containing missing parameter info
    """
    if not missing_params_summary:
        return
        
    # Create summary DataFrame
    summary_df = pd.DataFrame(missing_params_summary)
    summary_df.to_excel(writer, sheet_name='Missing Parameters Summary', index=False)
    
    # Get and format worksheet
    worksheet = writer.sheets['Missing Parameters Summary']
    
    # Auto-adjust columns
    for idx, col in enumerate(summary_df.columns, start=1):
        max_length = max(summary_df[col].astype(str).apply(len).max(), len(col))
        worksheet.column_dimensions[get_column_letter(idx)].width = max_length + 2
    
    # Color entire rows based on Missing Parameters
    for row in range(2, len(summary_df) + 2):  # Start from 2 to skip header
        missing_params = summary_df.iloc[row-2]['Missing Parameters']
        fill_color = green_fill if pd.isna(missing_params) or missing_params == '' else red_fill
        
        # Apply color to each cell in the row
        for col in range(1, len(summary_df.columns) + 1):
            cell = worksheet[f'{get_column_letter(col)}{row}']
            cell.fill = fill_color

# Import dei dati
refernece_file_path = 'data/reference.xlsx'
export_file_path = 'data/export.xlsx'
# Process all dataframes and save to single Excel file
df_reference = import_reference_data(refernece_file_path)
test_results = import_test_data(export_file_path)
df_creati = test_results['dataframes']

os.makedirs('results', exist_ok=True)
filename = os.path.join('results', 'infrarail_assets_checks.xlsx')
#initializing empty list to store missing parameters it has to be here since it is used populated by check_df_parameters and used by create_summary_sheet
missing_params_summary = []
# Colors for formatting for filling True/False values
green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')

# Usage in main code:
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    for df in df_creati:
        try:
            result = check_df_parameters(df_reference, df, writer, missing_params_summary)
            print(result)
        except Exception as e:
            print(f"Error processing dataframe: {str(e)}")
    
    create_summary_sheet(writer, missing_params_summary)

print(f"\nResults saved to: {filename}")